import sqlite3
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DB_NAME = "ventas_argos.db"

def generar_reporte():
    print("📊 Iniciando generación de reporte...")
    if os.path.exists(DB_NAME):
        db_path = DB_NAME
    elif os.path.exists(f"src/{DB_NAME}"):
        db_path = f"src/{DB_NAME}"
    else:
        print(f"❌ Error: No encuentro el archivo '{DB_NAME}'.")
        print("   Asegúrate de haber realizado al menos una venta.")
        return

    try:
     
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM ventas", conn)
        conn.close()

        if df.empty:
            print("⚠️ La base de datos existe pero está vacía. ¡Vende algo primero!")
            return

       
        df = df.rename(columns={
            "id": "ID Venta",
            "fecha": "Fecha y Hora",
            "cliente": "Nombre Cliente",
            "producto": "Detalle del Carrito", 
            "peso": "Categoría",
            "precio": "Monto Total (S/)",      
            "metodo_pago": "Medio de Pago",
            "direccion": "Dirección de Entrega"
        })

      
        fecha_hoy = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_excel = f"Reporte_Ventas_{fecha_hoy}.xlsx"

        df.to_excel(nombre_excel, index=False, engine='openpyxl')

        
        wb = load_workbook(nombre_excel)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(nombre_excel)
        
        print(f"✅ ¡LISTO! Reporte generado: {nombre_excel}")
        print(f"💰 Total recaudado en este reporte: S/ {df['Monto Total (S/)'].sum():.2f}")

    except Exception as e:
        print(f"❌ Ocurrió un error inesperado: {e}")

if __name__ == "__main__":
    generar_reporte()